import pandas as pd

input_columns = [1, 3]

data_empleados = pd.read_excel('./data/Control_de_acceso_docentes.xlsx  ',usecols=input_columns)

data_empleados = data_empleados.rename(columns={
    'Unnamed: 1': 'Nombre empleado',
    'Unnamed: 3': 'horarios'
})

# Eliminar primeras filas
data_empleados = data_empleados.iloc[5:].reset_index(drop=True)

# Eliminar filas sin Nombre empleado
data_empleados = data_empleados.dropna(subset=["Nombre empleado"])

# Limpiar espacios
data_empleados["Nombre empleado"] = (
    data_empleados["Nombre empleado"]
    .astype(str)
    .str.strip()
)

# Eliminar encabezados repetidos
data_empleados = data_empleados[
    ~data_empleados["Nombre empleado"].str.contains("Fecha", na=False)
]

# Identificar filas que son empleados
# Ejemplo: "Álzate Montoya Sandra Elena - 43743409"
data_empleados["es_empleado"] = data_empleados["Nombre empleado"].str.contains(
    r"\s-\s*",
    regex=True,
    na=False
)

# Guardar empleado actual y rellenar hacia abajo
data_empleados["empleado_completo"] = data_empleados["Nombre empleado"].where(
    data_empleados["es_empleado"]
).ffill()

# Separar nombre y documento
data_empleados[["nombre", "documento"]] = (
    data_empleados["empleado_completo"]
    .str.split(" - ", n=1, expand=True)
)

 

# Las filas que NO son empleado son las fechas
data_empleados["fecha"] = data_empleados["Nombre empleado"].where(
    ~data_empleados["es_empleado"]
)


 
data_empleados['horarios'] = data_empleados["horarios"] .str.split(" - ") .str[0]
 
# Dejar solo filas de fechas
resultado = data_empleados[
    ~data_empleados["es_empleado"]
].copy()

# Seleccionar columnas finales
resultado = resultado[[
    "nombre",
    "documento",
    "fecha",
    "horarios"
]]






# Exportar a Excel
ruta_salida = "./accesos_limpios_docentes.xlsx"
resultado.to_excel(ruta_salida, index=False)

import pandas as pd

# =========================
# LEER ARCHIVO FINAL
# =========================
df = pd.read_excel("./accesos_limpios_docentes.xlsx")

# =========================
# LIMPIAR HORARIOS
# =========================
df["horarios"] = (
    df["horarios"]
    .astype(str)
    .str.strip()
)

# Reemplazar textos vacíos por NaN
df["horarios"] = df["horarios"].replace(
    ["", "nan", "None"],
    pd.NA
)

# =========================
# CONVERTIR HORAS
# =========================
df["hora"] = pd.to_datetime(
    df["horarios"],
    format="%H:%M",
    errors="coerce"
)

# Hora límite
hora_limite = pd.to_datetime("06:21", format="%H:%M")

# =========================
# CREAR CONDICIONES
# =========================

# Llegada tarde
df["llegada_tarde"] = df["hora"] > hora_limite

# Llegada a tiempo
df["llegada_tiempo"] = df["hora"] <= hora_limite

# Sin marcar
df["sin_marcar"] = df["hora"].isna()

# =========================
# AGRUPAR POR EMPLEADO
# =========================
reporte = (
    df.groupby(["nombre", "documento"])
    .agg(
        llegadas_tarde=("llegada_tarde", "sum"),
        llegadas_a_tiempo=("llegada_tiempo", "sum"),
        sin_marcar=("sin_marcar", "sum")
    )
    .reset_index()
)

# =========================
# CALCULAR CUMPLIMIENTO
# =========================

reporte["total_registros"] = (
    reporte["llegadas_tarde"]
    + reporte["llegadas_a_tiempo"]
    + reporte["sin_marcar"]
)

reporte["cumplimiento"] = (
    (
        reporte["llegadas_a_tiempo"]
        / reporte["total_registros"]
    ) * 100
).round(2)

# =========================
# TOP TARDANZAS
# =========================
top_tardanzas = reporte.sort_values(
    by="llegadas_tarde",
    ascending=False
)

# =========================
# EXPORTAR
# =========================
# =========================
# EXPORTAR SOLO TABLA
# =========================

with pd.ExcelWriter(
    "./reportellegadasDocentes.xlsx",
    engine="openpyxl"
) as writer:

    reporte.to_excel(
        writer,
        sheet_name="Reporte General",
        index=False
    )

    # Ajustar ancho columnas
    worksheet = writer.sheets["Reporte General"]

    columnas = {
        "A": 50,
        "B": 16,
        "C": 18,
        "D": 22,
        "E": 15,
        "F": 18,
        "G": 18
    }

    for col, width in columnas.items():
        worksheet.column_dimensions[col].width = width
        
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =========================
# KPI DINÁMICO
# =========================
total_tardanzas = int(reporte["llegadas_tarde"].sum())
promedio_cumplimiento = round(reporte["cumplimiento"].mean(), 2)
total_sin_marcar = int(reporte["sin_marcar"].sum())

# Insertar filas arriba para las tarjetas
worksheet.insert_rows(1, 4)

# Tarjeta 1
worksheet.merge_cells("A1:B3")
cell = worksheet["A1"]
cell.value = f"TOTAL LLEGADAS TARDE\n{total_tardanzas}"
cell.font = Font(size=16, bold=True, color="FFFFFF")
cell.fill = PatternFill("solid", fgColor="1A74A8")
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Tarjeta 2
worksheet.merge_cells("C1:D3")
cell = worksheet["C1"]
cell.value = f"CUMPLIMIENTO PROMEDIO\n{promedio_cumplimiento}%"
cell.font = Font(size=16, bold=True, color="000000")
cell.fill = PatternFill("solid", fgColor="FCFF4C")
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Tarjeta 3
worksheet.merge_cells("E1:F3")
cell = worksheet["E1"]
cell.value = f"SIN MARCAR\n{total_sin_marcar}"
cell.font = Font(size=16, bold=True, color="FFFFFF")
cell.fill = PatternFill("solid", fgColor="000000")
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Bordes
borde = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for rango in ["A1:B3", "C1:D3", "E1:F3"]:
    for row in worksheet[rango]:
        for celda in row:
            celda.border = borde

# Alto de filas para que parezca tarjeta
worksheet.row_dimensions[1].height = 25
worksheet.row_dimensions[2].height = 25
worksheet.row_dimensions[3].height = 25