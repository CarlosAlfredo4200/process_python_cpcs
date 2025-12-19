import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------------------------------------------------
# 1️⃣ CONFIGURACIÓN GENERAL
# --------------------------------------------------------------------
input_columns = [0, 2, 5, 7, 16, 18]

ruta_pre_jardin = "./data/DATA_PRE-JARDIN_JARDIN_TRANSICION/a22757e746e74f0897c5f3230b6299d4.xlsx"
ruta_primero_once = "./data/DATA_1_11/a6242413b0694a429f74cbeb91c17a5d.xlsx"

# --------------------------------------------------------------------
# 2️⃣ FUNCIÓN DE LIMPIEZA (REUTILIZABLE)
# --------------------------------------------------------------------
def procesar_archivo(ruta_excel):
    data = pd.read_excel(ruta_excel, usecols=input_columns)

    # Eliminar filas iniciales basura
    df = data.iloc[16:].reset_index(drop=True)

    # Renombrar columnas
    df = df.rename(columns={
        "Unnamed: 0": "Estudiante",
        "Unnamed: 2": "Acudiente",
        "Unnamed: 5": "Grado",
        "Unnamed: 7": "Concepto",
        "Unnamed: 16": "Abono",
        "Unnamed: 18": "Estado",
    })

    # Limpiar nombres
    df["Estudiante"] = (
        df["Estudiante"].astype(str)
        .str.split("\n").str[0]
        .str.strip()
    )

    df["Acudiente"] = (
        df["Acudiente"].astype(str)
        .str.split("\n").str[0]
        .str.strip()
    )

    # Extraer año
    df["Año"] = df["Grado"].astype(str).str.extract(r"(20\d{2})", expand=False)

    # Extraer grado limpio
    df["Grado"] = (
        df["Grado"].astype(str)
        .str.extract(r"([A-Z\-ÁÉÍÓÚÑ]+)", expand=False)
    )

    # Rellenar hacia abajo
    df = df.ffill()

    # Limpiar registros inválidos
    df["Estudiante"] = df["Estudiante"].replace(["nan", "None", ""], pd.NA)
    df = df.dropna(subset=["Estudiante"])

    # ❌ Excluir encabezados y textos informativos
    patrones_excluir = (
        r"^Principal\s*-|"
        r"^Sede\s*-|"
        r"^Estudiante$|"
        r"^www\.q10\.com|"
        r"^Total\s+Principal"
    )

    df = df[
        ~df["Estudiante"].str.contains(
            patrones_excluir,
            case=False,
            regex=True,
            na=False
        )
    ]

    # Filtrar solo año 2026
    df = df[df["Año"] == "2026"]

    return df

# --------------------------------------------------------------------
# 3️⃣ PROCESAR AMBOS ARCHIVOS
# --------------------------------------------------------------------
df_pre_jardin = procesar_archivo(ruta_pre_jardin)
df_primero_once = procesar_archivo(ruta_primero_once)

# --------------------------------------------------------------------
# 4️⃣ UNIR TODA LA DATA
# --------------------------------------------------------------------
data_ed = pd.concat([df_pre_jardin, df_primero_once], ignore_index=True)

# --------------------------------------------------------------------
# 5️⃣ PAGADOS
# --------------------------------------------------------------------
data_ed_PAGADOS = data_ed[data_ed["Estado"] == "PAGADO"]

def pagados(grado):
    return data_ed_PAGADOS[data_ed_PAGADOS["Grado"] == grado]

# --------------------------------------------------------------------
# 6️⃣ EXPORTAR BASE
# --------------------------------------------------------------------
ruta_salida = "./base_matriculados_2026.xlsx"
data_ed.to_excel(ruta_salida, index=False)
print("\n✔ Archivo generado:", ruta_salida)

# --------------------------------------------------------------------
# 7️⃣ ESTILOS Y RESUMEN EN EXCEL
# --------------------------------------------------------------------
wb = load_workbook(ruta_salida)
ws = wb.active

# Estilo encabezados
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Auto ancho columnas
for column in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    ws.column_dimensions[column[0].column_letter].width = max_len + 2

# --------------------------------------------------------------------
# 8️⃣ RESUMEN FINAL
# --------------------------------------------------------------------
fila = ws.max_row + 2

totales = {
    "TOTAL GENERAL": len(data_ed),
    "TOTAL PAGADOS": len(data_ed_PAGADOS),
    "PAGADOS PRE-JARDIN": len(pagados("PRE-JARDIN")),
    "PAGADOS JARDIN": len(pagados("JARDIN")),
    "PAGADOS TRANSICION": len(pagados("TRANSICIÓN")),
    "PAGADOS PRIMERO": len(pagados("PRIMERO")),
    "PAGADOS SEGUNDO": len(pagados("SEGUNDO")),
    "PAGADOS TERCERO": len(pagados("TERCERO")),
    "PAGADOS CUARTO": len(pagados("CUARTO")),
    "PAGADOS QUINTO": len(pagados("QUINTO")),
    "PAGADOS SEXTO": len(pagados("SEXTO")),
    "PAGADOS SEPTIMO": len(pagados("SEPTIMO")),
    "PAGADOS OCTAVO": len(pagados("OCTAVO")),
    "PAGADOS NOVENO": len(pagados("NOVENO")),
    "PAGADOS DECIMO": len(pagados("DÉCIMO")),
    "PAGADOS ONCE": len(pagados("ONCE")),
}

for key, value in totales.items():
    ws[f"A{fila}"] = key
    ws[f"A{fila}"].font = Font(bold=True, size=11)
    ws[f"B{fila}"] = value
    ws[f"B{fila}"].font = Font(bold=True, size=12, color="007300")
    fila += 1

# Nota final
nota = (
    "Nota: 18 son deudores morosos con más de 4 pensiones. "
    "12 son de fundación y 20 sin reserva de cupo."
)

ws[f"A{fila+1}"] = nota
ws[f"A{fila+1}"].font = Font(bold=True, italic=True, size=11, color="000080")
ws.merge_cells(f"A{fila+1}:E{fila+1}")
ws[f"A{fila+1}"].alignment = Alignment(horizontal="left")

wb.save(ruta_salida)

print("\n📄 Archivo FINAL unificado (PRE-JARDÍN a ONCE) generado correctamente.\n")
